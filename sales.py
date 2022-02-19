import json
import pandas as pd
import numpy as np
import time
import datetime
import os
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import openpyxl
import cloudscraper

scraper = cloudscraper.create_scraper()

platform = "ps"  # Xbox = xone,    Playstation = ps,   PC = pc
directory = "2"  # Change to directory you want to use (category, ex: icons, heros, silvers)
want_pdf = 0 # Change to 1 if you want pdf with graphs, and 0 for no pdf
txt = "playerIDs.txt"  # Don't need to change
exc = "playerPrices.xlsx"  # Don't need to change
pdf = "SaleGraphs.pdf" # Don't need to change

filereference = open(directory+"/"+txt, "r")

# Make players directory and go through file to append the playersto the directory
players = dict()
print("Starting up...")
for line in filereference:
    line_stripped = line.strip()
    line_split = line_stripped.split(":")
    player = line_split[0].split("'")
    ID = line_split[1].split(",")
    ID = ID[0].strip()
    ID = int(ID)
    players[player[1]] = ID


# Cuts off 2% of lowest sales and 2% of highest sales so the deviation is gone
# Calculates the lowest, average, highest sales.
def calculations(pricelist):
    pricelist_copy = pricelist.copy() 
    prices = sorted(pricelist_copy, reverse=False)
    percents = int(round(len(pricelist) * 0.02))
    liste = prices[percents:-percents]
    low, high = min(liste), max(liste)
    average = np.mean(liste)
    return low, int(average), high


# Returns if the price is on uptrend or downtrend
def trend(pricelist):
    lengde = int(len(pricelist)//2)
    part1 = pricelist[:lengde]
    part2 = pricelist[lengde:]
    avg1 = np.mean(part1)
    avg2 = np.mean(part2)
    return str(round(((avg2-avg1)/avg1)*100, 1))


# Find buyprice (currently set buyprice to 10% lower than average, 5% profit after tax)
def buyprices(average, high, fluct, low):
    avg_avg = (average + (high*1000))/2
    buy = average*(0.9+fluct)
    if buy > average*0.95:
        buy = average*0.9
    if average >= 200000:
        buy = int((avg_avg*0.95 - 10000))
    if average >= 400000:
        buy = int((avg_avg*0.95 - 15000))
    if average >= 600000:
        buy = int((avg_avg*0.95 - 20000))
    if buy < int(low) and int(low) < average*0.95:
        buy = int(low)
    if buy < int(low) and int(low) > average*0.95:
        buy = int(low*0.95)
    return int(buy//1000)


# Most sales happen in these 5 percent
def most_sales_interval(avg, pricelist):
    count = 0
    commons = 0
    if avg < 100000:
        percennn = int((avg*0.025)//250)*250
        step = 250
    else:
        percennn = int((avg*0.025)//1000)*1000
        step = 1000
    if percennn < 1000:
        percennn = 1000
    if (max(pricelist)-min(pricelist)) < percennn*2:
        percennn = int(percennn//1.75)
    for i in range((min(pricelist)+percennn),(max(pricelist)-percennn),step):
        internalcount = 0
        for j in range(-percennn,percennn+1,step):
            internalcount += pricelist.count(i + j)
        if internalcount > count:
            count = internalcount
            commons = i
    interval = (int((commons-percennn)//1000),int((commons+percennn)//1000))
    return interval, count

# Percentage of sales over a given number
def sales_over_number(pricelist, number):
    new_list = []
    for price in pricelist:
        if price >= number:
            new_list.append(price)
    compare = (len(new_list) / len(pricelist))*100
    compare = round(compare,1)
    return str(compare)

# Median per three
def split_in_three(pricelist):
    lengde = len(pricelist)//3
    pt1 = np.median(pricelist[:lengde])
    pt2 = np.median(pricelist[lengde:lengde*2])
    pt3 = np.median(pricelist[-lengde:])
    return (int(pt1),int(pt2),int(pt3))

def date_to_string(date, dates):
    for i in range(0, 6):
        dater = date + datetime.timedelta(hours=i)
        year = dater.strftime("%Y")
        month = dater.strftime("%b")
        nr = dater.strftime("%d")
        hr = dater.strftime("%H")
        dato = str(month) + " " + str(nr) + " " + str(year[2:] + ", " + str(hr))
        for i in range(0, len(dates)):
            if dato in dates[i]:
                return i, (dates[i])[:-2]
    return 0, (dates[0])[:-2]

def timedifference(date):
    if date.find("21,") != -1:
        date = date.replace("21,", "2021")
    if date.find("22,") != -1:
        date = date.replace("22,", "2022")
    if date.find("pm") != -1:
        date = date.replace(" pm","")
    if date.find("am") != -1:
        date = date.replace(" am","")
    conve = datetime.datetime.strptime(date, "%b %d %Y %H:%M")
    now = datetime.datetime.now()
    diff = now - conve
    differnce_hrs = str(diff.days)+ " day(s) & "+ str(int(diff.seconds/3600)) + " hrs"
    return differnce_hrs, conve, diff

def plotgraph(x,date,title, mean, buyprice):
    fig = plt.figure(figsize=(12,5))
    plt.ylabel("Price")
    xlabel = "Last " + str(date)
    plt.xlabel(xlabel)
    y_avg = [mean]*len(x)
    y_buy = [buyprice*1000]*len(x)
    plt.title(title, fontsize=20)
    plt.plot(x, "b-", alpha=0.15,linewidth=1, label="Sales")
    plt.plot(x, "b.", markersize=4)
    avg_label = "Average: "+str(int(mean))
    plt.plot(y_avg, "m",label=avg_label, linestyle="--")
    buy_label = "Buyprice: "+str(buyprice)
    plt.plot(y_buy, "g",label=buy_label, linestyle="--")
    plt.legend(loc=2)
    y =[]
    for i in range(len(x)):
        y.append(i)
    plt.fill_between(y,x, alpha=0.15)
    if min(x) > buyprice*1000:
        btm = int((buyprice*1000)//1.05)
    else:
        btm = int(min(x)//1.05)
    plt.ylim(bottom=btm)
    plt.xlim(0,y[-1])
    plt.close()
    return fig

database = {}
playercount = 0
def find_player_data(liste):
    low_price, avg_price, high_price = calculations(liste)
    sorted_list = sorted(liste)
    lengde = int(len(liste)*0.9)
    potato = sorted_list[lengde]
    fluctuation = ((potato-avg_price)/avg_price)/2.2
    price_trend = trend(liste) + "%"
    best_sale_interval = most_sales_interval(avg_price, liste)
    buyprice = buyprices(avg_price, best_sale_interval[0][1], fluctuation, low_price)
    sales_over_avg = sales_over_number(liste, best_sale_interval[0][1]*1000)+"%"
    three = split_in_three(liste)
    return low_price, avg_price, high_price, price_trend, buyprice, best_sale_interval, sales_over_avg, three

if want_pdf == 1:
    pp = PdfPages(directory+"/"+pdf)
for (name, ID) in players.items():
    # Don't change this, you can be IP banned from futbin if you send too many requests within a time limit.
    time.sleep(1)
    tempdata = {}  # Temporary data

    # Gets player's sale data
    link = 'https://www.futbin.com/getPlayerChart?type=live-sales&resourceId=' + \
        str(ID) + '&platform=' + platform
    r_sell = scraper.get(link).text
    r_data = json.loads(r_sell)

    sales_list = []  # List of all sale prices, used for calculations
    dates = []
    for data in r_data:
        sales_list.append(data[1])
        dates.append(data[0])
    date = r_data[0][0]  # First date in sales list
    _, _, diff = timedifference(date)
    delta = 8
    if diff.days >= 3:
        delta = 36
    elif diff.days >= 1 and diff.days < 3:
        delta = 14
    else:
        delta = 8
    _, converted, _ = timedifference(r_data[-1][0])
    xhoursago = converted - datetime.timedelta(hours=delta)
    index, time2 = date_to_string(xhoursago,dates)
    time1,_,_ = timedifference(dates[index])
    sales_list = sales_list[index:]

    low_price, avg_price, high_price, price_trend, buyprice, best_sale_interval, sales_over_avg, three = find_player_data(sales_list)

    if want_pdf == 1:
        plot2 = plotgraph(sales_list, time1, str(name),avg_price,buyprice)
        pp.savefig(plot2)

    # Updates temporary data
    tempdata.update({"Buyprice": buyprice, "Lowest": low_price, "Average": avg_price,
                     "Highest": high_price,"Trend": price_trend, "Most sales int": best_sale_interval,
                     "Occurency rate": sales_over_avg, "Median per 1/3": three, "Data from last": time1, "# of sales": len(sales_list)})
    database.update({name: tempdata})  # Updates database

    # Prints the progress, how many players have been processed
    print(str(playercount + 1) + "/" + str(len(players)), end=": ")
    print(name)
    playercount += 1

# Write data to excel
date = str(datetime.datetime.now().date())
ExcelWorkbook = openpyxl.load_workbook(directory+"/"+exc)
writer = pd.ExcelWriter(directory+"/"+exc, engine='openpyxl')
worksheets = ExcelWorkbook.sheetnames
while len(worksheets) > 3:
    rmv = ExcelWorkbook[worksheets[0]]
    ExcelWorkbook.remove(rmv)
    worksheets = ExcelWorkbook.sheetnames
writer.book = ExcelWorkbook
df = pd.DataFrame(database).T
df.to_excel(writer, sheet_name=date)
writer.save()
if want_pdf == 1:
    pp.close()
print("Done, opening excel file...")
command = "open "+directory+"/"+exc
os.system(command)
if want_pdf == 1:
    command2 = "open "+directory+"/"+pdf
    os.system(command2)